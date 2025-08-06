import streamlit as st
from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook
import re

FUND_FILE = "table/income_data.xlsx"
AR_LOOKUP_FILE = "table/ar_code.xlsx"
SPEND_LOOKUP_FILE = "table/unique_spend_code.csv"
FUNDING_SOURCE_FILE = "table/funding_source.xlsx"
FISCAL_YEAR_FILE = "table/fiscal_year.xlsx"


if st.session_state.get("reset_flag", False):
    st.session_state.clear()
    st.session_state["just_reset"] = True
    st.rerun()

@st.cache_data(ttl=60)
def load_ar_lookup():
    if os.path.exists(AR_LOOKUP_FILE):
        df = pd.read_excel(AR_LOOKUP_FILE, dtype=str).fillna("")
        df.columns = df.columns.str.strip().str.replace("\ufeff", "", regex=False)
        df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
        return df
    return pd.DataFrame(columns=["รหัสโครงการวิจัย", "ar_code", "รหัสค่าใช้จ่าย"])
@st.cache_data
def load_spend_lookup():
    if os.path.exists(SPEND_LOOKUP_FILE):
        return pd.read_csv(SPEND_LOOKUP_FILE, dtype=str).fillna("")
    return pd.DataFrame(columns=["รหัสค่าใช้จ่าย", "หมวดรายจ่าย", "รายการ", "ประเภทค่าใช้จ่าย"])

@st.cache_data
def load_funding_source_data():
    if os.path.exists(FUNDING_SOURCE_FILE):
        return pd.read_excel(FUNDING_SOURCE_FILE, dtype=str).fillna("")
    return pd.DataFrame(columns=["รหัสงบประมาณ"])

@st.cache_data
def load_fiscal_year_data():
    if os.path.exists(FISCAL_YEAR_FILE):
        return pd.read_excel(FISCAL_YEAR_FILE, dtype=str).fillna("")
    return pd.DataFrame(columns=["ปีงบประมาณ"])

def lookup_spend_detail(spend_code):
    df = load_spend_lookup()
    matched = df[df["รหัสค่าใช้จ่าย"] == spend_code.strip()]
    if not matched.empty:
        return matched.iloc[0]["หมวดรายจ่าย"], matched.iloc[0]["รายการ"], matched.iloc[0]["ประเภทค่าใช้จ่าย"]
    return "", "", ""

def save_to_excel(df, filename):
    df.to_excel(filename, index=False)
    wb = load_workbook(filename)
    ws = wb.active
    for cell in ws['A']:
        cell.number_format = 'YYYY-MM-DD HH:mm:ss'
    wb.save(filename)

# --- Session State Setup ---
def init_session():
    st.session_state.setdefault("rounds", [0])
    st.session_state.setdefault("round_0_codes", [0])
    st.session_state.setdefault("fund_project_code", "")
    st.session_state.setdefault("fund_source", "")
    st.session_state.setdefault("fiscal_year", "")
    st.session_state.setdefault("contract_code", "")
    st.session_state.setdefault("duration_months", "")
    st.session_state.setdefault("fund_type", "")
    st.session_state.setdefault("fund_date", datetime.today())
    st.session_state.setdefault("contract_date", datetime.today())



if "rounds" not in st.session_state:
    init_session()

def add_round():
    new_idx = max(st.session_state.rounds) + 1
    st.session_state.rounds.append(new_idx)
    st.session_state[f"round_{new_idx}_codes"] = [0]

def remove_round():
    if len(st.session_state.rounds) > 1:
        last = st.session_state.rounds.pop()
        st.session_state.pop(f"round_{last}_codes", None)

def add_code(r_idx):
    st.session_state[f"round_{r_idx}_codes"].append(
        max(st.session_state[f"round_{r_idx}_codes"]) + 1
    )

def remove_code(r_idx, c_idx):
    key = f"round_{r_idx}_codes"
    if len(st.session_state[key]) > 1 and c_idx != 0:
        st.session_state[key].remove(c_idx)
st.markdown("""
    <style>
    .title-style {
        font-size:30px;
        font-weight:bold;
        color: #4a4a4a;
    }
    .section-header {
        font-size:20px;
        font-weight:bold;
        margin-top:20px;
        margin-bottom:10px;
        color: #333;
    }
    .block {
        border: 1px solid #ddd;
        border-radius: 10px;
        padding: 15px;
        margin-bottom: 10px;
        background-color: #fafafa;
    }
    </style>
""", unsafe_allow_html=True)

st.set_page_config(page_title="ระบบบันทึกทุนโครงการวิจัย", layout="wide")

# 🌟 Title
st.markdown('<div class="title-style">📋 ระบบบันทึกทุนโครงการวิจัย</div>', unsafe_allow_html=True)
st.write("")
# --- Main App UI ---
ar_lookup = load_ar_lookup()
spend_lookup = load_spend_lookup()
fund_source_df = load_funding_source_data()
fund_type_list = ["","ทุนภายใน", "ทุนภายนอก"]
fund_source_list1 = sorted(fund_source_df["รหัสงบประมาณ"].dropna().unique().tolist())
fund_source_list2 = [""] + fund_source_list1  # หรือ ["กรุณาเลือก"] + fund_source_list
fiscal_year_df = load_fiscal_year_data()
fiscal_year_list1 = sorted(fiscal_year_df["ปีงบประมาณ"].dropna().unique().tolist())
fiscal_year_list2 = [""] + fiscal_year_list1  # หรือ ["กรุณาเลือก"] + fund_source_list
st.set_page_config(page_title="ระบบบันทึกทุนโครงการวิจัย", layout="wide")
st.title("ระบบบันทึกทุนโครงการวิจัย")

# ตรวจสอบว่าค่าที่อยู่ใน session_state มีอยู่ใน list หรือไม่ เพื่อป้องกัน error
if "fund_source" not in st.session_state or st.session_state.fund_source not in fund_source_list2:
    st.session_state.fund_source = fund_source_list2[0] if fund_source_list2 else ""

# ตรวจสอบว่าค่าที่อยู่ใน session_state มีอยู่ใน list หรือไม่ เพื่อป้องกัน error
if "fiscal_year" not in st.session_state or st.session_state.fund_source not in fiscal_year_list2:
    st.session_state.fiscal_year = fiscal_year_list2[0] if fiscal_year_list2 else ""

# ตรวจสอบค่าหลัง reset และตั้งค่า default ถ้าไม่อยู่ใน list
if st.session_state.get("fund_source", "") not in fund_source_list2:
    st.session_state.fund_source = fund_source_list2[0] if fund_source_list2 else ""

if st.session_state.get("fiscal_year", "") not in fiscal_year_list2:
    st.session_state.fiscal_year = fiscal_year_list2[0] if fiscal_year_list2 else ""


st.session_state.fund_date = st.date_input("วันที่กรอกข้อมูล", value=st.session_state.fund_date)
st.session_state.fiscal_year = st.selectbox("ปีงบประมาณ",fiscal_year_list2,index=fiscal_year_list2.index(st.session_state.fiscal_year))
st.session_state.fund_project_code = st.text_input("รหัสโครงการวิจัย", value=st.session_state.fund_project_code).strip().upper()
pattern_fund_project_code = r"^E\d{4}_\d{3}$"
if not re.match(pattern_fund_project_code, st.session_state.fund_project_code):
    st.warning("รหัสโครงการวิจัยต้องอยู่ในรูปแบบ EXXXX_XXX (ตัวอย่าง: E2568_001)")
st.session_state.fund_type = st.selectbox("ประเภททุน", fund_type_list, index=fund_type_list.index(st.session_state.fund_type))
st.session_state.fund_source = st.selectbox("รหัสงบประมาณ",fund_source_list2,index=fund_source_list2.index(st.session_state.fund_source))
st.session_state.contract_date = st.date_input("วันที่เซ็นสัญญา", value=st.session_state.contract_date)
# st.session_state.duration_months = st.number_input("ระยะเวลาดำเนินโครงการ (เดือน)", min_value=1, step=1, value=st.session_state.duration_months)
duration_str = st.text_input("ระยะเวลาดำเนินโครงการ (เดือน)", value=str(st.session_state.get("duration_months", "")))
try:
    duration_val = int(duration_str)
except ValueError:
    duration_val = None

if duration_val is None or duration_val <= 0:
    st.warning("กรุณากรอกตัวเลขจำนวนเต็มที่มากกว่า 0 ในช่องระยะเวลาดำเนินโครงการ")
else:
    st.session_state.duration_months = duration_val

st.session_state.contract_code = st.text_input("รหัสสัญญา", value=st.session_state.contract_code)
pattern_contract_code = r"^CHR\d{3}/\d{4}$"
if not re.match(pattern_contract_code, st.session_state.contract_code):
    st.warning("รหัสสัญญาต้องอยู่ในรูปแบบ CHRXXX/XXXX (ตัวอย่าง: CHR001/2568)")


fund_project_code = st.session_state.fund_project_code
project_ar_codes = ar_lookup[ar_lookup["รหัสโครงการวิจัย"].str.upper() == fund_project_code]["ar_code"].unique().tolist()
has_ar = len(project_ar_codes) > 0

# --- รอบ (งวด)
for r_idx in st.session_state.rounds:
    with st.expander(f"📦 งวดที่ {r_idx+1}", expanded=True):
        st.markdown(f"### รายละเอียดงวดที่ {r_idx+1}")

        key_codes = f"round_{r_idx}_codes"
        if key_codes not in st.session_state:
            st.session_state[key_codes] = [0]

        if has_ar:
            # 1. ส่วน ar_code ตามเดิม
            ar_selected_list = st.multiselect(f"🔗 เลือก AR code สำหรับงวดที่ {r_idx+1}", project_ar_codes, key=f"ar_{r_idx}_multi")
            for ar_idx, ar_selected in enumerate(ar_selected_list):
                st.markdown(f'#### 🎯 AR code: {ar_selected}')
                rows = ar_lookup[
                    (ar_lookup["รหัสโครงการวิจัย"] == fund_project_code) &
                    (ar_lookup["ar_code"] == ar_selected)
                ]
                for i, row in rows.iterrows():
                    spend = row["รหัสค่าใช้จ่าย"]
                    cat, item, cost_type = lookup_spend_detail(spend)

                    col1, col2 = st.columns([2, 3])
                    with col1:
                        st.text_input("🔢 รหัสค่าใช้จ่าย", value=spend, key=f"code_{r_idx}_{ar_idx}_{i}", disabled=True)
                    with col2:
                        st.text_input("📂 หมวดรายจ่าย", value=cat, key=f"cat_{r_idx}_{ar_idx}_{i}", disabled=True)
                        st.text_input("📌 รายการ", value=item, key=f"item_{r_idx}_{ar_idx}_{i}", disabled=True)
                        st.text_input("🧾 ประเภทค่าใช้จ่าย", value=cost_type, key=f"cost_{r_idx}_{ar_idx}_{i}", disabled=True)
                        st.number_input("💰 จำนวนเงิน", min_value=0.0, step=100.0, key=f"amt_{r_idx}_{ar_idx}_{i}")

            # 2. ส่วนเพิ่มรหัสค่าใช้จ่ายนอก ar code (เหมือนตอนไม่มี ar code)
            st.markdown("#### ➕ เพิ่มรหัสค่าใช้จ่ายนอกกลุ่ม AR code")
            for c_idx in st.session_state[key_codes]:
                code_col, detail_col = st.columns([2, 5])
                with code_col:
                    code = st.text_input(f"🔢 รหัสค่าใช้จ่าย (งวด {r_idx+1} รายการ {c_idx+1})", key=f"round_{r_idx}_code_{c_idx}")
                cat, item, cost_type = lookup_spend_detail(code)
                with detail_col:
                    col1, col2 = st.columns(2)
                    with col1:
                        st.text_input("📂 หมวดรายจ่าย", value=cat, key=f"cat_{r_idx}_{c_idx}")
                        st.text_input("📌 รายการ", value=item, key=f"item_{r_idx}_{c_idx}")
                        st.text_input("🧾 ประเภทค่าใช้จ่าย", value=cost_type, key=f"cost_{r_idx}_{c_idx}")
                        st.number_input("💰 จำนวนเงิน", min_value=0.0, step=100.0, key=f"amt_free_{r_idx}_{c_idx}")
                if c_idx != 0:
                    if st.button(f"➖ ลบรายการ (งวด {r_idx+1} รายการ {c_idx+1})", key=f"btn_remove_{r_idx}_{c_idx}"):
                        remove_code(r_idx, c_idx)
                        st.rerun()
            if st.button(f"➕ เพิ่มรหัสค่าใช้จ่าย (งวด {r_idx+1})", key=f"btn_add_code_{r_idx}"):
                add_code(r_idx)
                st.rerun()

        else:
            # กรณีไม่มี ar code เหมือนเดิม
            for c_idx in st.session_state[key_codes]:
                code_col, detail_col = st.columns([2, 5])
                with code_col:
                    code = st.text_input(f"🔢 รหัสค่าใช้จ่าย (งวด {r_idx+1} รายการ {c_idx+1})", key=f"round_{r_idx}_code_{c_idx}")
                cat, item, cost_type = lookup_spend_detail(code)
                with detail_col:
                    col1, col2 = st.columns(2)
                    with col1:
                        st.text_input("📂 หมวดรายจ่าย", value=cat, key=f"cat_{r_idx}_{c_idx}")
                        st.text_input("📌 รายการ", value=item, key=f"item_{r_idx}_{c_idx}")
                        st.text_input("🧾 ประเภทค่าใช้จ่าย", value=cost_type, key=f"cost_{r_idx}_{c_idx}")
                        st.number_input("💰 จำนวนเงิน", min_value=0.0, step=100.0, key=f"amt_{r_idx}_{c_idx}")
                if c_idx != 0:
                    if st.button(f"➖ ลบรายการ (งวด {r_idx+1} รายการ {c_idx+1})", key=f"btn_remove_{r_idx}_{c_idx}"):
                        remove_code(r_idx, c_idx)
                        st.rerun()
            if st.button(f"➕ เพิ่มรหัสค่าใช้จ่าย (งวด {r_idx+1})", key=f"btn_add_code_{r_idx}"):
                add_code(r_idx)
                st.rerun()
        # รวมยอดแต่ละงวด
        total_amt = 0.0
        if has_ar:
            # ยอดจาก ar_code
            ar_selected_list = st.session_state.get(f"ar_{r_idx}_multi", [])
            for ar_idx, ar_code in enumerate(ar_selected_list):
                rows = ar_lookup[
                    (ar_lookup["รหัสโครงการวิจัย"] == fund_project_code) &
                    (ar_lookup["ar_code"] == ar_code)
                ]
                for idx, _ in rows.iterrows():
                    total_amt += st.session_state.get(f"amt_{r_idx}_{ar_idx}_{idx}", 0.0)

            # ยอดจากรหัสค่าใช้จ่ายอิสระนอกกลุ่ม ar_code
            for c_idx in st.session_state.get(f"round_{r_idx}_codes", [0]):
                total_amt += st.session_state.get(f"amt_free_{r_idx}_{c_idx}", 0.0)

        else:
            # กรณีไม่มี ar_code
            for c_idx in st.session_state.get(f"round_{r_idx}_codes", [0]):
                total_amt += st.session_state.get(f"amt_{r_idx}_{c_idx}", 0.0)
        st.info(f"💵 ยอดรวมงวดที่ {r_idx+1}: {total_amt:,.2f} บาท")

# ปุ่มเพิ่ม/ลบงวด
cols = st.columns([8, 1, 1])
cols[1].button("➕ เพิ่มงวด", on_click=add_round, key="btn_add_round")
if len(st.session_state.rounds) > 1:
    cols[2].button("➖ ลบงวด", on_click=remove_round, key="btn_remove_round")

# --- Reset function ---
def reset_form():
    tmp_rows = st.session_state.get("__tmp_new_rows__", None)

    st.session_state.clear()  # เคลียร์ทุกอย่าง
    
    if tmp_rows is not None:
        st.session_state["__tmp_new_rows__"] = tmp_rows  # เก็บไว้ต่อ

    st.session_state["just_reset"] = True
    st.rerun()

total_all = 0.0
for r_idx in st.session_state.rounds:
    if has_ar:
        # ยอดจาก ar_code
        ar_selected_list = st.session_state.get(f"ar_{r_idx}_multi", [])
        for ar_idx, ar_code in enumerate(ar_selected_list):
            rows = ar_lookup[
                (ar_lookup["รหัสโครงการวิจัย"] == fund_project_code) &
                (ar_lookup["ar_code"] == ar_code)
            ]
            for idx, _ in rows.iterrows():
                total_all += st.session_state.get(f"amt_{r_idx}_{ar_idx}_{idx}", 0.0)

        # ยอดจากรหัสค่าใช้จ่ายอิสระนอกกลุ่ม ar_code
        for c_idx in st.session_state.get(f"round_{r_idx}_codes", [0]):
            total_all += st.session_state.get(f"amt_free_{r_idx}_{c_idx}", 0.0)
    else:
        for c_idx in st.session_state.get(f"round_{r_idx}_codes", [0]):
            total_all += st.session_state.get(f"amt_{r_idx}_{c_idx}", 0.0)

st.markdown("---")
st.markdown(f"## 💰 ยอดรวมทั้งหมด: {total_all:,.2f} บาท")


form_valid  = (
    st.session_state.fiscal_year != "" and
    st.session_state.fund_source != "" and
    st.session_state.fund_type != "" and
    st.session_state.fund_project_code != "" and
    st.session_state.contract_code != "" and
    st.session_state.rounds != "" and
    duration_str != "" and
    re.match(pattern_fund_project_code, st.session_state.fund_project_code) and
    re.match(pattern_contract_code, st.session_state.contract_code)
)

if not form_valid:
    st.error("⚠️ กรุณากรอกข้อมูลให้ครบถ้วนและถูกต้องก่อนบันทึก")

# ปุ่มบันทึก (จะ disabled ถ้าไม่ valid)
if st.button("💾 บันทึกข้อมูลทุนทั้งหมด", disabled=not form_valid, key="btn_save_all"):
    all_rows = []
    for r_idx in st.session_state.rounds:
        round_num = r_idx + 1
        if has_ar:
            # บันทึกรายการจาก ar_code ตามเดิม
            ar_selected_list = st.session_state.get(f"ar_{r_idx}_multi", [])
            for ar_idx, ar_code in enumerate(ar_selected_list):
                rows = ar_lookup[
                    (ar_lookup["รหัสโครงการวิจัย"] == fund_project_code) &
                    (ar_lookup["ar_code"] == ar_code)
                ]
                for idx, row in rows.iterrows():
                    spend = row["รหัสค่าใช้จ่าย"]
                    cat, item, cost_type = lookup_spend_detail(spend)
                    amt = st.session_state.get(f"amt_{r_idx}_{ar_idx}_{idx}", 0.0)
                    all_rows.append({
                        "วันที่กรอกข้อมูล": datetime.combine(st.session_state.fund_date, datetime.now().time()),
                        "ปีงบประมาณ" : st.session_state.fiscal_year,
                        "รหัสโครงการวิจัย": fund_project_code,
                        "ประเภททุน": st.session_state.fund_type,
                        "รหัสงบประมาณ" : st.session_state.fund_source,
                        "วันที่เซนสัญญา": st.session_state.contract_date,
                        "ระยะเวลาดำเนินโครงการ (เดือน)": st.session_state.duration_months,
                        "รหัสสัญญา": st.session_state.contract_code,
                        "งวด": round_num,
                        "ar_code": ar_code,
                        "รหัสค่าใช้จ่าย": spend,
                        "หมวดรายจ่าย": cat,
                        "รายการ": item,
                        "ประเภทค่าใช้จ่าย": cost_type,
                        "จำนวนเงิน": amt
                    })
            # **เพิ่มบันทึกรหัสค่าใช้จ่ายนอกกลุ่ม AR code ด้วย**
            for c_idx in st.session_state.get(f"round_{r_idx}_codes", [0]):
                code = st.session_state.get(f"round_{r_idx}_code_{c_idx}", "").strip()
                if code == "":
                    continue
                cat = st.session_state.get(f"cat_{r_idx}_{c_idx}", "")
                item = st.session_state.get(f"item_{r_idx}_{c_idx}", "")
                cost_type = st.session_state.get(f"cost_{r_idx}_{c_idx}", "")
                amt = st.session_state.get(f"amt_free_{r_idx}_{c_idx}", 0.0)
                all_rows.append({
                    "วันที่กรอกข้อมูล": datetime.combine(st.session_state.fund_date, datetime.now().time()),
                    "ปีงบประมาณ" : st.session_state.fiscal_year,
                    "รหัสโครงการวิจัย": fund_project_code,
                    "ประเภททุน": st.session_state.fund_type,
                    "รหัสงบประมาณ" : st.session_state.fund_source,
                    "วันที่เซนสัญญา": st.session_state.contract_date,
                    "ระยะเวลาดำเนินโครงการ (เดือน)": st.session_state.duration_months,
                    "รหัสสัญญา": st.session_state.contract_code,
                    "งวด": round_num,
                    "ar_code": "",  # ไม่มี ar_code สำหรับรหัสนอกกลุ่ม
                    "รหัสค่าใช้จ่าย": code,
                    "หมวดรายจ่าย": cat,
                    "รายการ": item,
                    "ประเภทค่าใช้จ่าย": cost_type,
                    "จำนวนเงิน": amt
                })
        else:
            # กรณีไม่มี ar_code เหมือนเดิม
            for c_idx in st.session_state.get(f"round_{r_idx}_codes", [0]):
                code = st.session_state.get(f"round_{r_idx}_code_{c_idx}", "").strip()
                if code == "":
                    continue
                cat = st.session_state.get(f"cat_{r_idx}_{c_idx}", "")
                item = st.session_state.get(f"item_{r_idx}_{c_idx}", "")
                cost_type = st.session_state.get(f"cost_{r_idx}_{c_idx}", "")
                amt = st.session_state.get(f"amt_{r_idx}_{c_idx}", 0.0)
                all_rows.append({
                    "วันที่กรอกข้อมูล": datetime.combine(st.session_state.fund_date, datetime.now().time()),
                    "ปีงบประมาณ" : st.session_state.fiscal_year,
                    "รหัสโครงการวิจัย": fund_project_code,
                    "ประเภททุน": st.session_state.fund_type,
                    "รหัสงบประมาณ" : st.session_state.fund_source,
                    "วันที่เซนสัญญา": st.session_state.contract_date,
                    "ระยะเวลาดำเนินโครงการ (เดือน)": st.session_state.duration_months,
                    "รหัสสัญญา": st.session_state.contract_code,
                    "งวด": round_num,
                    "ar_code": "",
                    "รหัสค่าใช้จ่าย": code,
                    "หมวดรายจ่าย": cat,
                    "รายการ": item,
                    "ประเภทค่าใช้จ่าย": cost_type,
                    "จำนวนเงิน": amt
                })

    if all_rows:
        df = pd.DataFrame(all_rows)
        if os.path.exists(FUND_FILE):
            old_df = pd.read_excel(FUND_FILE)
            df = pd.concat([old_df, df], ignore_index=True)

        save_to_excel(df, FUND_FILE)
        st.success("✅ บันทึกข้อมูลทุนทั้งหมดเรียบร้อยแล้ว")
        st.cache_data.clear()
        
        st.session_state["__tmp_new_rows__"] = all_rows
        reset_form()




# --- เมื่อรีเซตเสร็จ ให้คืนค่า new_rows กลับมา ---
if st.session_state.get("just_reset", False):
    if "__tmp_new_rows__" in st.session_state:
        st.session_state["new_rows"] = st.session_state["__tmp_new_rows__"]
        del st.session_state["__tmp_new_rows__"]

    st.success("🔄 ฟอร์มถูกรีเซตเรียบร้อยแล้ว")
    del st.session_state["just_reset"]


if "new_rows" in st.session_state:
    new_rows_df = pd.DataFrame(st.session_state["new_rows"])
    st.subheader("📋 ข้อมูลที่เพิ่งเพิ่มล่าสุด")
    st.dataframe(new_rows_df.astype(str))