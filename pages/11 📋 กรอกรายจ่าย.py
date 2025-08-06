import streamlit as st
from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook

FUND_FILE = "table/expend_data.xlsx"
AR_LOOKUP_FILE = "table/ar_code.xlsx"
SPEND_LOOKUP_FILE = "table/unique_spend_code.csv"
INCOME_FILE = "table/income_data.xlsx"

@st.cache_data(ttl=60)
def load_ar_lookup():
    if os.path.exists(AR_LOOKUP_FILE):
        df = pd.read_excel(AR_LOOKUP_FILE, dtype=str).fillna("")
        df.columns = df.columns.str.strip().str.replace("\ufeff", "", regex=False)
        df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))  # ✅ ใหม่
        return df
    return pd.DataFrame(columns=["รหัสโครงการวิจัย", "ar_code", "รหัสค่าใช้จ่าย"])

@st.cache_data
def load_income_data():
    if os.path.exists(INCOME_FILE):
        df = pd.read_excel(INCOME_FILE, dtype=str).fillna("")
        df.columns = df.columns.str.strip()
        return df
    return pd.DataFrame()

@st.cache_data
def load_spend_lookup():
    if os.path.exists(SPEND_LOOKUP_FILE):
        df = pd.read_csv(SPEND_LOOKUP_FILE, dtype=str).fillna("")
        df.columns = df.columns.str.strip()
        return df
    return pd.DataFrame(columns=["รหัสค่าใช้จ่าย", "หมวดรายจ่าย", "รายการ", "ประเภทค่าใช้จ่าย"])

def lookup_spend_detail(spend_code):
    df = load_spend_lookup()
    matched = df[df["รหัสค่าใช้จ่าย"] == spend_code.strip()]
    if not matched.empty:
        return (matched.iloc[0]["หมวดรายจ่าย"], matched.iloc[0]["รายการ"], matched.iloc[0]["ประเภทค่าใช้จ่าย"])
    return "", "", ""

def save_to_excel(df, filename):
    df.to_excel(filename, index=False)
    wb = load_workbook(filename)
    ws = wb.active
    for cell in ws['A']:
        cell.number_format = 'YYYY-MM-DD HH:mm:ss'
    wb.save(filename)

def init_session():
    st.session_state.fund_project_code = ""
    st.session_state.fund_type = ""
    st.session_state.fund_date = datetime.today()
    st.session_state.contract_date = datetime.today()
    st.session_state.contract_payment_type = ""
    st.session_state.contract_code = ""


def reset_form():
    tmp_rows = st.session_state.get("__tmp_new_rows__", None)

    st.session_state.clear()  # เคลียร์ทุกอย่าง
    
    if tmp_rows is not None:
        st.session_state["__tmp_new_rows__"] = tmp_rows  # เก็บไว้ต่อ

    st.session_state["just_reset"] = True
    st.rerun()



required_keys = ["fund_date", "contract_payment_type", "contract_code", "fund_project_code"]
if not all(key in st.session_state for key in required_keys):
    init_session()

st.set_page_config(page_title="ระบบบันทึกทุนโครงการวิจัย", layout="wide")
st.title("📋 ระบบบันทึกทุนโครงการวิจัย")

ar_lookup = load_ar_lookup()
income_df = load_income_data()

st.session_state.fund_date = st.date_input("📅 วันที่กรอกข้อมูล", value=st.session_state.fund_date)

fund_project_code_list1 = sorted(income_df["รหัสโครงการวิจัย"].dropna().unique().tolist())
fund_project_code_list2 = [""] + fund_project_code_list1  # หรือ ["กรุณาเลือก"] + fund_source_list
st.session_state.fund_project_code = st.selectbox("รหัสโครงการวิจัย", fund_project_code_list2, index=fund_project_code_list2.index(st.session_state.fund_project_code))

if st.session_state.fund_project_code:
    filtered_income = income_df[
        income_df["รหัสโครงการวิจัย"] == st.session_state.fund_project_code
    ]
    available_fund_type = sorted(filtered_income["ประเภททุน"].dropna().unique().tolist())
    st.session_state.fund_type = st.selectbox(
        "ประเภททุน",
        available_fund_type,
        index=available_fund_type.index(st.session_state.fund_type) if st.session_state.fund_type in available_fund_type else 0
    )

# ป้องกัน error โดยใช้ get()
contract_payment_type_list = ["","ค่าใช้จ่ายจริง", "เงินยืมทดรองจ่าย"]
st.session_state.contract_payment_type = st.selectbox("💼 ประเภทการจ่ายเงิน", contract_payment_type_list, index=contract_payment_type_list.index(st.session_state.contract_payment_type))
st.session_state.contract_date = st.date_input("📅 วันที่เบิกจ่าย", value=st.session_state.contract_date)

contract_code_input = st.text_input("🔢 รหัสกิจกรรม (13 หลัก)", value=st.session_state.contract_code)
if contract_code_input and (not contract_code_input.isdigit() or len(contract_code_input) != 13):
    st.error("❌ รหัสกิจกรรมต้องเป็นตัวเลข 13 หลักเท่านั้น")
    st.stop()
else:
    st.session_state.contract_code = contract_code_input

if st.session_state.fund_project_code:
    filtered_income = income_df[
        (income_df["รหัสโครงการวิจัย"] == st.session_state.fund_project_code) &
        (income_df["ประเภททุน"] == st.session_state.fund_type)
    ]
    available_rounds = sorted(filtered_income["งวด"].dropna().astype(int).unique().tolist())

    if not available_rounds:
        st.warning("❗ ไม่พบข้อมูลงวดในระบบสำหรับโครงการและประเภททุนนี้")
        st.stop()

    selected_rounds = st.multiselect("📦 เลือกงวดที่ต้องการดูข้อมูล", available_rounds)
    if not selected_rounds:
        st.info("👉 กรุณาเลือกงวดอย่างน้อย 1 งวด")
        st.stop()

    grand_total = 0.0

    for selected_round in selected_rounds:
        st.markdown(f"### 📦 ข้อมูลงวดที่ {selected_round}")

        round_income = filtered_income[filtered_income["งวด"].astype(int) == selected_round]
        valid_spend_codes = round_income["รหัสค่าใช้จ่าย"].dropna().unique().tolist()

        total_amt = 0.0
        data_rows = []

        # แสดงข้อมูลที่มี AR code
        ar_rows = ar_lookup[
            (ar_lookup["รหัสโครงการวิจัย"] == st.session_state.fund_project_code) &
            (ar_lookup["ar_code"].notna()) & (ar_lookup["ar_code"].str.strip() != "") &
            (ar_lookup["รหัสค่าใช้จ่าย"].isin(valid_spend_codes))
        ]

        for ar_code in ar_rows["ar_code"].unique():
            st.markdown(f'#### 🎯 AR code: {ar_code}')
            rows = ar_rows[ar_rows["ar_code"] == ar_code]
            for i, row in rows.iterrows():
                spend = row["รหัสค่าใช้จ่าย"]
                cat, item, cost_type = lookup_spend_detail(spend)
                col1, col2 = st.columns([2, 3])
                with col1:
                    st.text_input("🔢 รหัสค่าใช้จ่าย", value=spend, key=f"code_{selected_round}_{ar_code}_{i}", disabled=True)
                with col2:
                    st.text_input("📂 หมวดรายจ่าย", value=cat, key=f"cat_{selected_round}_{ar_code}_{i}", disabled=True)
                    st.text_input("📌 รายการ", value=item, key=f"item_{selected_round}_{ar_code}_{i}", disabled=True)
                    st.text_input("🧾 ประเภทค่าใช้จ่าย", value=cost_type, key=f"cost_{selected_round}_{ar_code}_{i}", disabled=True)
                    amt = st.number_input("💰 จำนวนเงิน", min_value=0.0, step=100.0, key=f"amt_{selected_round}_{ar_code}_{i}")
                    total_amt += amt
                    data_rows.append((selected_round, ar_code, spend, cat, item, cost_type, amt))

        # แสดงข้อมูลที่ไม่มี AR code (นอกกลุ่ม AR code)
        # หารหัสค่าใช้จ่ายทั้งหมดในรอบนี้
        spend_codes_in_round = set(valid_spend_codes)

        # หารหัสที่มี AR code ไปแล้ว
        ar_used_spend_codes = set(ar_rows["รหัสค่าใช้จ่าย"].unique())

        # รหัสที่ไม่มี AR code
        spend_codes_without_ar = spend_codes_in_round - ar_used_spend_codes

        if spend_codes_without_ar:
            st.markdown(f"### 🧾 รหัสค่าใช้จ่ายนอกกลุ่ม AR code")
            for i, spend in enumerate(spend_codes_without_ar):
                cat, item, cost_type = lookup_spend_detail(spend)
                col1, col2 = st.columns([2, 3])
                with col1:
                    st.text_input("🔢 รหัสค่าใช้จ่าย", value=spend, key=f"outside_code_{selected_round}_{i}", disabled=True)
                    st.text_input("📂 หมวดรายจ่าย", value=cat, key=f"outside_cat_{selected_round}_{i}", disabled=True)
                    st.text_input("📌 รายการ", value=item, key=f"outside_item_{selected_round}_{i}", disabled=True)
                    st.text_input("🧾 ประเภทค่าใช้จ่าย", value=cost_type, key=f"outside_cost_{selected_round}_{i}", disabled=True)
                with col2:
                    amt = st.number_input("💰 จำนวนเงิน", min_value=0.0, step=100.0, key=f"amt_free_{selected_round}_{i}")
                    total_amt += amt
                    data_rows.append((selected_round, "", spend, cat, item, cost_type, amt))

        grand_total += total_amt
        st.info(f"💵 ยอดรวมงวดที่ {selected_round}: {total_amt:,.2f} บาท")
        st.success(f"💰💰 ยอดรวมทั้งหมดของทุกงวด: {grand_total:,.2f} บาท")
        if st.button(f"💾 บันทึกข้อมูลงวด {selected_round}", key=f"btn_save_{selected_round}"):
            if data_rows:
                saved_rows = []
                for round_no, ar_code, spend, cat, item, cost_type, amt in data_rows:
                    if round_no == selected_round and amt > 0:
                        saved_rows.append({
                            "วันที่กรอกข้อมูล": datetime.combine(st.session_state.fund_date, datetime.now().time()),
                            "รหัสโครงการวิจัย": st.session_state.fund_project_code,
                            "ประเภททุน": st.session_state.fund_type,
                            "ประเภทการจ่ายเงิน": st.session_state.contract_payment_type,
                            "วันที่เบิกจ่าย": st.session_state.contract_date,
                            "รหัสกิจกรรม": st.session_state.contract_code,
                            "งวด": round_no,
                            "ar_code": ar_code,
                            "รหัสค่าใช้จ่าย": spend,
                            "หมวดรายจ่าย": cat,
                            "รายการ": item,
                            "ประเภทค่าใช้จ่าย": cost_type,
                            "จำนวนเงิน": amt
                        })

                df = pd.DataFrame(saved_rows)
                if os.path.exists(FUND_FILE):
                    old_df = pd.read_excel(FUND_FILE)
                    df = pd.concat([old_df, df], ignore_index=True)
                try:
                    save_to_excel(df, FUND_FILE)
                    st.success(f"✅ บันทึกข้อมูลเรียบร้อยแล้ว")
                    st.cache_data.clear()
                    st.session_state["__tmp_new_rows__"] = saved_rows
                    reset_form()  # ✅ เพิ่มบรรทัดนี้
                except Exception as e:
                    st.error(f"❌ ไม่สามารถบันทึก {FUND_FILE} ได้: {e}")

                # 🔄 เฉพาะกรณีเป็นเงินยืมทดรองจ่าย ให้บันทึกลง reserve_payment.xlsx ด้วย
                if st.session_state.contract_payment_type == "เงินยืมทดรองจ่าย" :
                    reserve_rows = []
                    for row in saved_rows:
                        borrow_date = pd.to_datetime(row["วันที่เบิกจ่าย"])
                        return_date = borrow_date + pd.Timedelta(days=90)
                        reserve_rows.append({
                            "วันที่กรอกข้อมูล": row["วันที่กรอกข้อมูล"],
                            "รหัสโครงการวิจัย": row["รหัสโครงการวิจัย"],
                            "ar_code": row["ar_code"],
                            "รหัสค่าใช้จ่าย": row["รหัสค่าใช้จ่าย"],
                            "วันที่ยืม": row["วันที่เบิกจ่าย"],
                            "จำนวนเงิน": row["จำนวนเงิน"],
                            "วันที่ต้องคืน": pd.to_datetime(row["วันที่เบิกจ่าย"]) + pd.Timedelta(days=90),
                            "วันที่คืนเงิน": return_date,
                            "เงินที่คืน": "",
                            "คงเหลือ": "",
                            "สถานะ": ""
                        })
                    reserve_df = pd.DataFrame(reserve_rows)
                    reserve_file = "table/reserve_payment.xlsx"

                    if os.path.exists(reserve_file):
                        old_reserve = pd.read_excel(reserve_file)
                        frames = [df for df in [old_reserve, reserve_df] if not df.empty and not df.isna().all().all()]
                        reserve_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

                    try:
                        save_to_excel(reserve_df, reserve_file)
                        st.info("📁 บันทึกข้อมูลเงินยืมทดรองจ่ายลง reserve_payment.xlsx แล้ว")
                        st.cache_data.clear()
                    except Exception as e:
                        st.error(f"❌ ไม่สามารถบันทึก reserve_payment.xlsx ได้: {e}")
            else:
                st.warning("⚠️ ไม่มีข้อมูลให้บันทึก")
                
# --- เมื่อรีเซตเสร็จ ให้คืนค่า new_rows กลับมา ---
if st.session_state.get("just_reset", False):
    if "__tmp_new_rows__" in st.session_state:
        st.session_state["new_rows"] = st.session_state["__tmp_new_rows__"]
        del st.session_state["__tmp_new_rows__"]

    st.success("🔄 ฟอร์มถูกรีเซตเรียบร้อยแล้ว")
    del st.session_state["just_reset"]


if "new_rows" in st.session_state:
    new_rows_df = pd.DataFrame(st.session_state["new_rows"])
    st.subheader("📋 ข้อมูลที่เพิ่งเพิ่มล่าสุด")
    st.dataframe(new_rows_df.astype(str))